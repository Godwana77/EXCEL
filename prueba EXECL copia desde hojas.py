#Copia los valores de Libro2 a Libro1


import openpyxl
from openpyxl.styles import PatternFill
#E:\Users\Mariano\Repositorios\Python\EXCEL\prueba EXECL copia desde hojas.py
def procesar_excel(libro_fuente, hoja_fuente, columna_buscar, libro_destino, hoja_destino, columna_datos, columna_resultado):
    # Cargar los libros y hojas de trabajo
    wb_fuente = openpyxl.load_workbook(libro_fuente)
    ws_fuente = wb_fuente[hoja_fuente]
    
    wb_destino = openpyxl.load_workbook(libro_destino)
    ws_destino = wb_destino[hoja_destino]

    # Crear un diccionario de datos del libro destino key= columna A. valores desde A a H
    datos_destino = {}
    for fila in ws_destino.iter_rows(min_row=2, values_only=True):
        clave = fila[columna_datos - 1]
        datos_destino[clave] = fila
      

    # Estilo para marcar las filas en azul claro

    
    fill_azul_claro = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Procesar la columna de búsqueda y copiar datos coincidentes
    for fila in range(2, ws_fuente.max_row + 1):
        valor_a_buscar = ws_fuente.cell(row=fila, column=columna_buscar).value
        
        # Solo tomar los últimos 17 caracteres
     
        
        if valor_a_buscar in datos_destino: # Si valor_a_buscar se encuentra en el diccionario
            resultado = datos_destino[valor_a_buscar]
            ws_fuente.cell(row=fila, column=columna_resultado, value=resultado[1])
            # Marcar la fila en azul claro
            for col in range(1, ws_fuente.max_column + 1):
                ws_fuente.cell(row=fila, column=col).fill = fill_azul_claro
              
                ws_destino[valor_a_buscar].fill = fill_azul_claro
        else:
            ws_fuente.cell(row=fila, column=columna_resultado, value="No encontrado")

    # Guardar los cambios en el libro fuente
    wb_fuente.save(libro_fuente)
    print(f"Se ha completado el procesamiento y se guardaron los resultados en '{libro_fuente}'.")

# Ejemplo de uso
procesar_excel(
    libro_fuente="libro1.xlsx", #Aqui se copian lo datos de Lbros2
    hoja_fuente="Hoja1",
    columna_buscar=1,   # Columna con datos a buscar (en números, A=1, B=2...)
    libro_destino="libro2.xlsx", #De aqui se extraen los datos
    hoja_destino="Hoja1",
    columna_datos=1,    # Columna con claves en el libro destino
    columna_resultado=10 # Columna donde copiar los resultados
    #Abria que cambiar para que coja las colmnas necesarias
    )