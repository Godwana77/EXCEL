import openpyxl
from openpyxl.styles import PatternFill


class ProcesadorExcel():
    def __init__(self, facturas, bd, columnas1, columnas2, columnas_destino):
        """
        Inicializa el procesador con los archivos y las columnas.
        :param archivo1: Nombre del primer archivo Excel.
        :param archivo2: Nombre del segundo archivo Excel.
        :param columnas1: Lista de columnas a buscar en el primer archivo.
        :param columnas2: Lista de columnas a buscar en el segundo archivo.
        :param columnas_destino: Lista de columnas destino en el primer archivo.
        :columna_encargo : Columna encargo
        """
        self.facturas = facturas
        self.bd = bd
        self.columnas1 = columnas1
        self.columnas2 = columnas2
        self.columnas_destino = columnas_destino
        self.FILL_AZUL = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        self.COLUMNA_ENCARGO="G"
        self.COLUMNA_FACTURADA="H"
        self.BD_IMP_HOJA="D"
        self.BD_IMP_FACT="E"
        self.FACT_IMP_IMP="D"
    def cargar_archivos(self):
        """Carga los archivos Excel en memoria."""
        self.wbfact = openpyxl.load_workbook(self.facturas)
        self.wbbd = openpyxl.load_workbook(self.bd)
        self.hoja1 = self.wbfact.active
        self.hoja2 = self.wbbd["Hojadestino"]

    def procesar(self):
        """Procesa las coincidencias entre las columnas de ambos archivos."""
        for fila1 in range(2, self.hoja1.max_row + 1):  # Asumiendo encabezado en fila 1
            for col1, col2, col_destino in zip(self.columnas1, self.columnas2, self.columnas_destino):
                valor1 = self.hoja1[f"{col1}{fila1}"].value
                if valor1 is None:
                    continue

                for fila2 in range(2, self.hoja2.max_row + 1):  # Asumiendo encabezado en fila 1
                    valor2 = self.hoja2[f"{col2}{fila2}"].value

                    if valor1 == valor2:  # Si coinciden
                        # Aqui se copian los valores
                        #hoja1=factura
                        # hoja 2 BD
                        
                        #Copia el encargo
                        self.hoja1[f"{col_destino}{fila1}"].value = self.hoja2[f"{self.COLUMNA_ENCARGO}{str(fila2)}"].value
                        # Marcar como facturada
                        self.hoja2[f"{self.COLUMNA_FACTURADA}{str(fila2)}"].value="F"
                        # Copiar el importe 
                        self.hoja2[f"{self.BD_IMP_FACT}{str(fila2)}"].value=self.hoja1[f"{self.FACT_IMP_IMP}{fila1}"].value
                        
                        # Marcar en azul ambas celdas
                        self.hoja1[f"{col1}{fila1}"].fill = self.FILL_AZUL
                        self.hoja2[f"{col2}{fila2}"].fill = self.FILL_AZUL

    def guardar_archivos(self):
        """Guarda los cambios en nuevos archivos con nombres descriptivos."""
       
        self.wbfact.save(facturas)
        self.wbfact.close()
        self.wbbd.save(bd)
        self.wbbd.close()
        
        print(f"Archivos guardados como '{self.facturas}' y '{self.bd}'.")

    def ejecutar(self):
        """Ejecuta el flujo completo de procesamiento."""
        self.cargar_archivos()
        self.procesar()
        self.guardar_archivos()

# Uso de la clase
if __name__ == "__main__":
    # Configuración
    
    facturas = "facturas.xlsx"  # Primer archivo
    bd = "BD.xlsx"  # Segundo archivo
    columnas1 = ["C"]      # Columnas a buscar en el primer archivo
    columnas2 = ["C"]      # Columnas a buscar en el segundo archivo
    columnas_destino = ["I", "J"]  # Columnas destino en el primer archivo

    # Crear instancia y ejecutar
    procesador = ProcesadorExcel(facturas, bd, columnas1, columnas2, columnas_destino)
    procesador.ejecutar()